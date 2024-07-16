from openpyxl.utils import get_column_letter, column_index_from_string
from openpyxl.utils.cell import coordinate_from_string
import re
import logging
from typing import Dict, List, Tuple, Set

# Set up logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

def is_blank_cell(cell) -> bool:
    """Check if a cell is truly blank (no value and no borders)."""
    return (
        cell.value is None and
        cell.border.left.style is None and
        cell.border.right.style is None and
        cell.border.top.style is None and
        cell.border.bottom.style is None
    )

def has_all_borders(cell) -> bool:
    """Check if a cell has borders on all sides."""
    return (
        cell.border.left.style is not None and
        cell.border.right.style is not None and
        cell.border.top.style is not None and
        cell.border.bottom.style is not None
    )

def compare_cell_coordinates(coord1: str, coord2: str) -> int:
    """Compare two cell coordinates."""
    try:
        match1 = re.match(r'([A-Z]+)([0-9]+)', coord1)
        match2 = re.match(r'([A-Z]+)([0-9]+)', coord2)
        if match1 and match2:
            col1, row1 = match1.groups()
            col2, row2 = match2.groups()
            if row1 != row2:
                return int(row1) - int(row2)
            return column_index_from_string(col1) - column_index_from_string(col2)
        return 0
    except ValueError as e:
        logger.error(f"Error comparing cell coordinates {coord1} and {coord2}: {e}")
        return 0  # Return 0 if comparison fails

def flood_fill_box(ws, start_row: int, start_col: int) -> Dict[str, str]:
    """Use flood fill algorithm to detect an info box."""
    box = {"start": f"{get_column_letter(start_col)}{start_row}", "end": f"{get_column_letter(start_col)}{start_row}"}
    stack = [(start_row, start_col)]
    visited = set()

    while stack:
        row, col = stack.pop()
        if (row, col) in visited:
            continue

        visited.add((row, col))
        cell = ws.cell(row=row, column=col)

        if is_blank_cell(cell):
            continue

        # Update box boundaries
        current_coord = f"{get_column_letter(col)}{row}"
        if compare_cell_coordinates(current_coord, box["start"]) < 0:
            box["start"] = current_coord
        if compare_cell_coordinates(current_coord, box["end"]) > 0:
            box["end"] = current_coord

        # Check neighboring cells
        for dr, dc in [(-1, 0), (1, 0), (0, -1), (0, 1)]:
            new_row, new_col = row + dr, col + dc
            if 1 <= new_row <= ws.max_row and 1 <= new_col <= ws.max_column:
                stack.append((new_row, new_col))

    return box

def parse_cell_reference(cell_ref: str) -> Tuple[str, int]:
    """Parse a cell reference into column and row, with error handling."""
    try:
        column, row = coordinate_from_string(cell_ref)
        return column, row
    except ValueError as e:
        logger.error(f"Error parsing cell reference {cell_ref}: {e}")
        return 'A', 1  # Return default values if parsing fails

def detect_info_boxes(ws, headers: List[Dict[str, str]]) -> List[Dict[str, any]]:
    """Detect info boxes in the worksheet and associate headers."""
    info_boxes = []
    visited: Set[Tuple[int, int]] = set()

    for row in range(1, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            if (row, col) not in visited and has_all_borders(ws.cell(row=row, column=col)):
                box = flood_fill_box(ws, row, col)
                start_col, start_row = parse_cell_reference(box["start"])
                end_col, end_row = parse_cell_reference(box["end"])
                
                try:
                    visited.update((r, c) for r in range(start_row, end_row + 1)
                                    for c in range(column_index_from_string(start_col),
                                                   column_index_from_string(end_col) + 1))
                except ValueError as e:
                    logger.error(f"Error updating visited cells for box {box['start']}:{box['end']}: {e}")
                    continue  # Skip this box if there's an error
                
                # Associate headers
                box_headers = []
                for header in headers:
                    header_cell = header["cell"]
                    if (compare_cell_coordinates(box["start"], header_cell) <= 0 and
                        compare_cell_coordinates(header_cell, box["end"]) <= 0):
                        box_headers.append(header)
                
                info_boxes.append({"range": f"{box['start']}:{box['end']}", "headers": box_headers})

    return info_boxes
