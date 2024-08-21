import openpyxl
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.utils import get_column_letter, column_index_from_string
import datetime

def list_ranges(filename, sheet=None):
    """
    Returns list of named ranges that are globally defined or sheet-specific.
    These are names that represent either a cell or range of cells.

    :param filename: Path to the Excel file
    :param sheet: Sheet name (optional). If provided, returns sheet-specific ranges
    :return: List of range names
    """
    try:
        workbook = openpyxl.load_workbook(filename, read_only=True, data_only=True)
    except FileNotFoundError:
        raise FileNotFoundError(f"File not found: {filename}")
    except InvalidFileException:
        raise ValueError(f"Invalid Excel file: {filename}")

    if sheet:
        if sheet not in workbook.sheetnames:
            raise ValueError(f"Sheet '{sheet}' not found in the workbook")
        return [name for name, range_obj in workbook.defined_names.items() if range_obj.localSheetId == workbook.sheetnames.index(sheet)]
    else:
        return [name for name, range_obj in workbook.defined_names.items() if range_obj.localSheetId is None]

def is_merged_cell(cell, merged_ranges):
    for merged_range in merged_ranges:
        if cell.coordinate in merged_range:
            return True, str(merged_range)
    return False, None

def read_ranges(filename, range_list, include_empty=False, include_merge_info=True):
    """
    Reads data from specified ranges in the Excel file.

    :param filename: Path to the Excel file
    :param range_list: List of range names to read
    :param include_empty: Whether to include empty cells in the output
    :param include_merge_info: Whether to include merge information
    :return: Dictionary with range names as keys and data as values
    """
    try:
        workbook = openpyxl.load_workbook(filename, read_only=False, data_only=False)
    except FileNotFoundError:
        raise FileNotFoundError(f"File not found: {filename}")
    except InvalidFileException:
        raise ValueError(f"Invalid Excel file: {filename}")

    result = {}
    for range_name in range_list:
        if range_name in workbook.defined_names:
            try:
                range_string = workbook.defined_names[range_name].attr_text
                print(f"Range string for {range_name}: {range_string}")

                # Handle invalid references
                if '#REF!' in range_string:
                    print(f"Warning: Invalid reference in range '{range_name}'")
                    result[range_name] = None
                    continue

                # Parse the range string
                if '!' in range_string:
                    sheet_name, cell_range = range_string.rsplit('!', 1)
                    sheet_name = sheet_name.strip("'")
                    if sheet_name.startswith('[') and sheet_name.endswith(']'):
                        sheet_name = sheet_name[1:-1]
                else:
                    sheet_name = workbook.active.title
                    cell_range = range_string

                if sheet_name in workbook.sheetnames:
                    ws = workbook[sheet_name]
                    merged_ranges = ws.merged_cells.ranges
                    cells = ws[cell_range]

                    result[range_name] = {
                        "data": [],
                        "merged_cells": [] if include_merge_info else None,
                        "repeated_values": []
                    }

                    if isinstance(cells, (openpyxl.cell.read_only.ReadOnlyCell, openpyxl.cell.cell.Cell)):
                        merged, merge_range = is_merged_cell(cells, merged_ranges)
                        cell_data = [cells.coordinate, cells.value if cells.value != '#DIV/0!' else None]
                        if include_empty or cell_data[1] is not None:
                            result[range_name]["data"].append(cell_data)
                        if merged and include_merge_info:
                            result[range_name]["merged_cells"].append(merge_range)
                    else:
                        value_count = {}
                        for row in cells:
                            row_data = []
                            for cell in row:
                                merged, merge_range = is_merged_cell(cell, merged_ranges)
                                cell_data = [cell.coordinate, cell.value if cell.value != '#DIV/0!' else None]
                                # make sure that cell_data is JSON serializable
                                # for example, datetime is not, so
                                # we convert it to string
                                if isinstance(cell_data[1], datetime.datetime):
                                    cell_data[1] = cell_data[1].isoformat()
                                else:
                                    cell_data[1] = cell_data[1]


                                if include_empty or cell_data[1] is not None:
                                    row_data.extend(cell_data)
                                if merged and include_merge_info and merge_range not in result[range_name]["merged_cells"]:
                                    result[range_name]["merged_cells"].append(merge_range)
                                if cell_data[1] is not None:
                                    value_count[cell_data[1]] = value_count.get(cell_data[1], 0) + 1
                            if row_data:
                                result[range_name]["data"].append(row_data)

                        # Process repeated values
                        for value, count in value_count.items():
                            if count > 1:
                                cells_with_value = [cell[0] for row in result[range_name]["data"] for cell in zip(row[::2], row[1::2]) if cell[1] == value]
                                result[range_name]["repeated_values"].append([value, cells_with_value])
                else:
                    print(f"Warning: Sheet '{sheet_name}' not found for range '{range_name}'")
                    result[range_name] = None
            except Exception as e:
                print(f"Error processing range '{range_name}': {str(e)}")
                result[range_name] = None
        else:
            result[range_name] = None  # Range not found

    return result

def write_range(filename, range_name, values):
    """
    Writes values to a specified range in the Excel file.
    Range name must be absolute including the worksheet name.

    :param filename: Path to the Excel file
    :param range_name: Name of the range to write to
    :param values: List of lists in row-by-row grid, may include values or formulas
    :return: Boolean indicating success
    """
    try:
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        raise FileNotFoundError(f"File not found: {filename}")
    except InvalidFileException:
        raise ValueError(f"Invalid Excel file: {filename}")

    if range_name not in workbook.defined_names:
        raise ValueError(f"Range '{range_name}' not found in the workbook")

    dests = workbook.defined_names[range_name].destinations
    for sheet, coord in dests:
        ws = workbook[sheet]
        cells = ws[coord]

        if isinstance(cells, openpyxl.cell.cell.Cell):
            if not isinstance(values, list) or len(values) != 1 or len(values[0]) != 2:
                raise ValueError("Input values do not match the range size (single cell)")
            cells.value = values[0][1]
        else:
            flat_values = [item for sublist in values for item in sublist]
            if len(flat_values) != len(cells) * 2:
                raise ValueError("Input values do not match the range size")
            for i, cell in enumerate(cells):
                cell.value = flat_values[i*2 + 1]

    try:
        workbook.save(filename)
        return True
    except PermissionError:
        raise PermissionError(f"Unable to save the file. It might be open in another application.")
    except Exception as e:
        raise Exception(f"An error occurred while saving the file: {str(e)}")

# Example usage:
if __name__ == "__main__":
    filename = "valuation.xlsx"
    print("Global ranges:", list_ranges(filename))

    # print out values in all global ranges
    #for name in list_ranges(filename):
    #    print(f"Reading '{name}':", read_ranges(filename, [name]))

    print("Reading 'ExpensesBox':", read_ranges(filename, ["ExpensesBox"]))
    #print("Writing to 'NamedRange2':", write_range(filename, "NamedRange2", [["A1", "=10*20"]]))
