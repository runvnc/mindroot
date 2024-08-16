import openpyxl
from openpyxl.utils import get_column_letter, range_boundaries
from typing import List, Union
from datetime import datetime, date

def process_cell_value(value):
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    elif isinstance(value, (int, float, str)):
        return value
    elif value is None:
        return None
    elif hasattr(value, 'value'):  # Handle openpyxl cell objects
        return process_cell_value(value.value)
    elif isinstance(value, openpyxl.worksheet.formula.ArrayFormula):
        return str(value)
    else:
        return str(value)

def process_cell(cell, cell_value, cell_formula, cell_ref):
    if cell_value is None and (cell_formula == '' or cell_formula is None):
        return []  # Return an empty list for empty cells
    processed_value = process_cell_value(cell_value)
    processed_formula = process_cell_value(cell_formula) if cell_formula else None
    return [
        cell_ref,
        processed_value,
        processed_formula
    ]

def excel_to_nested_lists(file_path: str, sheet_name: str, arrangement: str = 'row', cell_range: str = None) -> List[List[List[Union[str, None, int, float]]]]:
    if arrangement not in ['row', 'column']:
        raise ValueError("Arrangement must be either 'row' or 'column'")

    wb = openpyxl.load_workbook(file_path, data_only=False)
    wb_values = openpyxl.load_workbook(file_path, data_only=True)
    
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in the workbook.")
    
    ws = wb[sheet_name]
    ws_values = wb_values[sheet_name]
    
    if cell_range:
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    else:
        min_row, min_col = 1, 1
        max_row, max_col = ws.max_row, ws.max_column
    
    if arrangement == 'row':
        result = []
        for row in range(min_row, max_row + 1):
            row_data = []
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell_value = ws_values.cell(row=row, column=col).value
                cell_formula = cell.data_type == 'f' and cell.value or ''
                cell_ref = f"{get_column_letter(col)}{row}"
                cell_data = process_cell(cell, cell_value, cell_formula, cell_ref)
                row_data.append(cell_data)
            if row_data:  # Only append non-empty rows
                result.append(row_data)
    else:  # column-based arrangement
        result = []
        for col in range(min_col, max_col + 1):
            col_data = []
            for row in range(min_row, max_row + 1):
                cell = ws.cell(row=row, column=col)
                cell_value = ws_values.cell(row=row, column=col).value
                cell_formula = cell.data_type == 'f' and cell.value or ''
                cell_ref = f"{get_column_letter(col)}{row}"
                cell_data = process_cell(cell, cell_value, cell_formula, cell_ref)
                col_data.append(cell_data)
            if col_data:  # Only append non-empty columns
                result.append(col_data)
    
    return result


if __name__ == '__main__':
    excel_file = 'valuation.xlsx'
    sheet_name = 'Inputs'
    output_row_based = excel_to_nested_lists(excel_file, sheet_name, 'row')
    output_column_based = excel_to_nested_lists(excel_file, sheet_name, 'column')
    print(output_row_based)
    #print(output_column_based)

# Usage example:
# excel_file = 'path/to/your/spreadsheet.xlsx'
# sheet_name = 'Sheet1'
# output_row_based = excel_to_nested_lists(excel_file, sheet_name, 'row', 'A1:C10')
# output_column_based = excel_to_nested_lists(excel_file, sheet_name, 'column', 'A1:C10')
# print(output_row_based)
# print(output_column_based)
