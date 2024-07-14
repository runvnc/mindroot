import asyncio
from mod import open_workbook, select_sheet
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side

class MockContext:
    def __init__(self):
        self.data = {}

async def create_test_workbook(filename):
    wb = Workbook()
    ws = wb.active
    ws.title = "Test Sheet"

    # Create headers
    headers = ["ID", "Name", "Age", "Salary"]
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Add some data
    data = [
        (1, "John Doe", 30, 50000),
        (2, "Jane Smith", 28, 55000),
        (3, "Bob Johnson", 35, 60000)
    ]
    for row, record in enumerate(data, start=2):
        for col, value in enumerate(record, start=1):
            ws.cell(row=row, column=col, value=value)

    # Create an info box
    ws.merge_cells('F2:G3')
    info_box = ws['F2']
    info_box.value = "Company Info"
    info_box.font = Font(bold=True)
    info_box.alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    for row in ws['F2:G3']:
        for cell in row:
            cell.border = thin_border

    # Add some empty rows and columns
    ws.insert_rows(5)
    ws.insert_cols(5)

    wb.save(filename)

async def test_excel_analysis():
    test_file = "test_workbook.xlsx"
    await create_test_workbook(test_file)

    context = MockContext()
    
    # Open the workbook
    result = await open_workbook(test_file, context)
    print(result)

    # Select and analyze the sheet
    result = await select_sheet("Test Sheet", context)
    print(result)

    # Print the full structure
    print("\nFull structure:")
    print(context.data['sheet_structure'])

if __name__ == "__main__":
    asyncio.run(test_excel_analysis())
