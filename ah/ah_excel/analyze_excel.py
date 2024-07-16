from openpyxl.styles import Font, Alignment, Border, Fill
from openpyxl.utils import get_column_letter, column_index_from_string
from collections import defaultdict
import traceback
from .info_box_detection import detect_info_boxes

def is_header(cell, row, merged_ranges):
    # Only consider cells in the first 5 rows as potential headers
    if row > 5:
        return False, None
    
    # Check if the cell contains a formula
    if cell.data_type == 'f':
        return False, None
    
    # Check if the cell contains only numbers
    if isinstance(cell.value, (int, float)):
        return False, None
    
    # Count the number of header-like characteristics
    header_score = 0
    
    if cell.font.bold:
        header_score += 1
    
    if cell.alignment.horizontal == 'center':
        header_score += 1
    
    if cell.fill.start_color.index != '00000000':  # Has fill color
        header_score += 1
    
    if isinstance(cell.value, str):
        if cell.value.isupper() or len(cell.value.split()) <= 3:
            header_score += 1
    
    # Check if the cell is part of a merged range
    merged_range = next((str(rng) for rng in merged_ranges if cell.coordinate in rng), None)
    
    # Consider it a header if it meets at least 2 criteria
    return header_score >= 2, merged_range

def analyze_structure(ws):
    """Analyze the structure of the given worksheet."""
    try:
        structure = {
            "headers": [],
            "data_ranges": [],
            "empty_columns": [],
            "empty_rows": [],
            "merged_cells": [],
            "info_boxes": [],
            "text_cells": [],
            "numeric_cells": [],
        }

        # Get all merged cell ranges
        merged_ranges = ws.merged_cells.ranges

        # Analyze cells
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                cell = ws.cell(row=row, column=col)
                if cell.value is not None:
                    is_header_cell, merged_range = is_header(cell, row, merged_ranges)
                    if is_header_cell:
                        header_info = {
                            "cell": cell.coordinate,
                            "value": str(cell.value)
                        }
                        if merged_range:
                            header_info["merged_range"] = merged_range
                        structure["headers"].append(header_info)
                    elif isinstance(cell.value, (int, float)):
                        structure["numeric_cells"].append(cell.coordinate)
                    else:
                        structure["text_cells"].append(cell.coordinate)

        # Analyze data ranges and empty columns/rows
        data_ranges = defaultdict(lambda: {"start": None, "end": None})
        for col in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col)
            data_found = False
            for row in range(1, ws.max_row + 1):
                if ws.cell(row=row, column=col).value is not None:
                    data_found = True
                    if data_ranges[col_letter]["start"] is None:
                        data_ranges[col_letter]["start"] = row
                    data_ranges[col_letter]["end"] = row
            
            if not data_found:
                structure["empty_columns"].append(col_letter)
            elif data_ranges[col_letter]["start"] is not None:
                structure["data_ranges"].append(f"{col_letter}{data_ranges[col_letter]['start']}:{col_letter}{data_ranges[col_letter]['end']}")

        # Detect empty rows
        for row in range(1, ws.max_row + 1):
            if all(ws.cell(row=row, column=col).value is None for col in range(1, ws.max_column + 1)):
                structure["empty_rows"].append(row)

        # Detect merged cells
        structure["merged_cells"] = [str(merged_cell) for merged_cell in merged_ranges]

        # Detect info boxes
        structure["info_boxes"] = detect_info_boxes(ws, structure["headers"])

        # Limit the number of cells reported to prevent excessive output
        max_cells = 1000
        if len(structure["text_cells"]) > max_cells:
            structure["text_cells"] = structure["text_cells"][:max_cells]
            structure["text_cells"].append(f"... and {len(structure['text_cells']) - max_cells} more")
        if len(structure["numeric_cells"]) > max_cells:
            structure["numeric_cells"] = structure["numeric_cells"][:max_cells]
            structure["numeric_cells"].append(f"... and {len(structure['numeric_cells']) - max_cells} more")

        return structure
    except Exception as e:
        # print full stack trace
        traceback.print_exc()
        return {"error": str(e)}
