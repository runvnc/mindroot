from ..commands import command
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter, column_index_from_string, range_boundaries
from collections import defaultdict
import json
from .analyze_excel import analyze_structure
from .output_data import excel_to_nested_lists

from .excel_recalculator import recalculate_excel


@command()
async def list_sheets(filename, context):
    """Open an Excel workbook and return its sheets.
    Always use this before other commands unless you are sure you know the spreadsheet structure.
    Example:
    { "list_sheets": { "filename": "example.xlsx" } }
    """
    wb = load_workbook(filename)
    return f"Opened workbook {filename}. Sheets: {', '.join(wb.sheetnames)}"

@command()
async def analyze_sheet(filename, sheet_name, context=None):
    """Analyze the structure of a specific sheet.
    Example:
    { "analyze_sheet": { "filename": "example.xlsx", "sheet_name": "Sheet1" } }
    """
    try:
        wb = load_workbook(filename)
        if sheet_name not in wb.sheetnames:
            return f"Sheet {sheet_name} not found in the workbook."
        
        structure = analyze_structure(wb[sheet_name])
        
        summary = {
            "headers": structure["headers"],
            "info_boxes": structure["info_boxes"],
            "data_ranges_count": len(structure["data_ranges"]),
            "empty_columns_count": len(structure["empty_columns"]),
            "empty_rows_count": len(structure["empty_rows"]),
            "merged_cells_count": len(structure["merged_cells"]),
            "text_cells_count": len(structure["text_cells"]),
            "numeric_cells_count": len(structure["numeric_cells"]),
        }
        
        return summary 
    except Exception as e:
        return f"Error: {str(e)}"

@command()
async def read_cells(filename, sheet_name, arrangement='row', context=None):
    """Read all cells from sheet and return as nested lists.
    Example:
    { "read_cells": { "filename": "example.xlsx", "sheet_name": "Sheet1", "arrangement": "row" } }
    """
    try:
        result = excel_to_nested_lists(filename, sheet_name, arrangement)
        rows = len(result)
        cols = len(result[0])
        max_allowed_cells = 400
        total_cells = rows * cols
        max_row = max(rows, total_cells // cols)
        result = result[:max_row]
        if rows > max_row:
            result.append([f"Output truncated to {max_row} rows. Total rows: {rows}"])
        return json.dumps(result)
    except Exception as e:
        return f"Error: {str(e)}"

@command()
async def write_cell(filename, sheet_name, cell_reference, value, context=None):
    """Write a value to a specific cell.
    Example:
    { "write_cell": { "filename": "example.xlsx", "sheet_name": "Sheet1", "cell_reference": "C7", "value": 1500 } }
    """
    try:
        wb = load_workbook(filename)
        ws = wb[sheet_name]
        ws[cell_reference] = value
        wb.save(filename)
        recalculate_excel(filename)
        return f"Wrote {value} to cell {cell_reference} in {filename}, sheet {sheet_name}."
    except Exception as e:
        return f"Error: {str(e)}"

@command()
async def write_cell_range(filename, sheet_name, cell_range, values, overwrite_formulas=False, context=None):
    """Write values to a rectangular range of cells, in row order.
       IMPORTANT: check the location of formulas, and consider using insert_rows first if the your data would otherwise
       overwrite formulas.

       cell_range - is specified as upperLeft:lowerRight cell

       values - nested array of values. these will typically be numbers (specified as numbers, not strings!),
                but could also be strings or formulas (if overwrite_formulas is true).

       overwrite_formulas - typically it important to specify False, otherwise built-in calculations will be broken.

    Example ( note cells are specified in order as [ [ A1, B1, C1], [ A2, B2, C2 ], [ A3, B3, C3 ] ] ):

    { "write_cell_range": { "filename": "example.xlsx", "sheet_name": "Sheet1", "cell_range": "A1:C3", "values": [[1, 2, 3], [4, 5, 6], [7, 8, 9]],
                            "overwrite_formulas": false } }
    """
    try:
        wb = load_workbook(filename)
        ws = wb[sheet_name]
        
        min_col, min_row, max_col, max_row = range_boundaries(cell_range)
        
        if len(values) != max_row - min_row + 1 or any(len(row) != max_col - min_col + 1 for row in values):
            return "Error: Input shape does not match the specified range."


        # Check for existing formulas in the range
        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell_ref = f"{get_column_letter(col)}{row}"
                if ws[cell_ref].data_type == "f" and not overwrite_formulas:
                    return f"Error: Cell {cell_ref} contains a formula and overwrite_formulas is set to False."

        for row in range(min_row, max_row + 1):
            for col in range(min_col, max_col + 1):
                cell_ref = f"{get_column_letter(col)}{row}"
                value = values[row - min_row][col - min_col]
                ws[cell_ref] = value
        
        wb.save(filename)
        recalculate_excel(filename)
        return f"Updated range {cell_range} in {filename}, sheet {sheet_name} successfully."
    except Exception as e:
        return f"Error: {str(e)}"

@command()
async def insert_rows(filename, sheet_name, row_number, num_rows, context=None):
    """Insert a specified number of rows before a given row number.
    Example:
    { "insert_rows": { "filename": "/the/path/example.xlsx", "sheet_name": "Sheet1", "row_number": 5, "num_rows": 3 } }
    """
    try:
        wb = load_workbook(filename)
        ws = wb[sheet_name]
        ws.insert_rows(row_number, amount=num_rows)
        wb.save(filename)
        recalculate_excel(filename)
        return f"Inserted {num_rows} rows before row {row_number} in {filename}, sheet {sheet_name}."
    except Exception as e:
        return f"Error: {str(e)}"

@command()
async def insert_columns(filename, sheet_name, column_letter, num_columns, context=None):
    """Insert a specified number of columns before a given column letter.
    Example:
    { "insert_columns": { "filename": "/the/path/example.xlsx", "sheet_name": "Sheet1", "column_letter": "C", "num_columns": 2 } }
    """
    try:
        wb = load_workbook(filename)
        ws = wb[sheet_name]
        column_index = column_index_from_string(column_letter)
        ws.insert_cols(column_index, amount=num_columns)
        wb.save(filename)
        recalculate_excel(filename)
        return f"Inserted {num_columns} columns before column {column_letter} in {filename}, sheet {sheet_name}."
    except Exception as e:
        return f"Error: {str(e)}"


