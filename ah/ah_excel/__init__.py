from ..commands import command
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from collections import defaultdict
import json
import io

# Module-level cache for the workbook
_workbook_cache = None

def _get_workbook():
    global _workbook_cache
    if _workbook_cache is None:
        raise ValueError("No workbook is currently cached. Use open_workbook first.")
    return load_workbook(filename=io.BytesIO(_workbook_cache))

def _save_workbook_cache(wb):
    global _workbook_cache
    buffer = io.BytesIO()
    wb.save(buffer)
    _workbook_cache = buffer.getvalue()

@command()
async def open_workbook(filename, context=None):
    """Open an Excel workbook and cache it in memory.
    Example:
    { "open_workbook": { "filename": "example.xlsx" } }
    """
    wb = load_workbook(filename)
    _save_workbook_cache(wb)
    context.data['current_workbook'] = filename
    context.data['current_sheet'] = wb.active.title
    return f"Opened and cached workbook {filename}. Sheets: {', '.join(wb.sheetnames)}"

@command()
async def select_sheet(sheet_name, context=None):
    """Select a specific sheet and analyze its structure.
    Example:
    { "select_sheet": { "sheet_name": "Sheet1" } }
    """
    wb = _get_workbook()
    if sheet_name not in wb.sheetnames:
        return f"Sheet {sheet_name} not found in the workbook."
    
    context.data['current_sheet'] = sheet_name
    structure = analyze_structure(wb[sheet_name])
    return f"Selected sheet {sheet_name}. Structure:\n{json.dumps(structure, indent=2)}"

@command()
async def read_cell(cell_reference, context=None):
    """Read the value of a specific cell.
    Example:
    { "read_cell": { "cell_reference": "B5" } }
    """
    wb = _get_workbook()
    ws = wb[context.data['current_sheet']]
    return str(ws[cell_reference].value)

@command()
async def write_cell(cell_reference, value, context=None):
    """Write a value to a specific cell.
    Example:
    { "write_cell": { "cell_reference": "C7", "value": 1500 } }
    """
    wb = _get_workbook()
    ws = wb[context.data['current_sheet']]
    ws[cell_reference] = value
    _save_workbook_cache(wb)
    return f"Wrote {value} to cell {cell_reference} and updated the cache."

@command()
async def save_workbook(filename=None, context=None):
    """Save the workbook to disk, optionally with a new filename.
    Example:
    { "save_workbook": { "filename": "updated_file.xlsx" } }
    """
    wb = _get_workbook()
    if filename is None:
        filename = context.data['current_workbook']
    wb.save(filename)
    context.data['current_workbook'] = filename
    return f"Saved workbook as {filename}"

def analyze_structure(ws):
    """Analyze the structure of the given worksheet."""
    structure = {
        "headers": [],
        "data_ranges": [],
        "empty_columns": [],
        "empty_rows": [],
        "merged_cells": [],
    }

    # Detect headers (assume first row with data)
    for col in range(1, ws.max_column + 1):
        cell_value = ws.cell(row=1, column=col).value
        if cell_value:
            structure["headers"].append({
                "column": get_column_letter(col),
                "value": str(cell_value)
            })

    # Analyze data ranges and empty columns/rows
    data_ranges = defaultdict(lambda: {"start": None, "end": None})
    for col in range(1, ws.max_column + 1):
        col_letter = get_column_letter(col)
        data_found = False
        for row in range(2, ws.max_row + 1):  # Start from row 2 to skip header
            if ws.cell(row=row, column=col).value is not None:
                data_found = True
                if data_ranges[col_letter]["start"] is None:
                    data_ranges[col_letter]["start"] = row
                data_ranges[col_letter]["end"] = row
        
        if not data_found:
            structure["empty_columns"].append(col_letter)
        elif data_ranges[col_letter]["start"] is not None:
            structure["data_ranges"].append(f"{col_letter}{data_ranges[col_letter]['start']}:{col_letter}{data_ranges[col_letter]['end']}")

    # Detect empty rows
    for row in range(1, ws.max_row + 1):
        if all(ws.cell(row=row, column=col).value is None for col in range(1, ws.max_column + 1)):
            structure["empty_rows"].append(row)

    # Detect merged cells
    structure["merged_cells"] = [str(merged_cell) for merged_cell in ws.merged_cells.ranges]

    return structure